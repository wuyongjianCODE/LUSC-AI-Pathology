import json
import re
import torch
import torch.nn as nn
import torch.optim as optim
import torchvision.models as models
import torchvision.transforms as transforms
import pandas as pd
import numpy as np
from PIL import Image
from pathlib import Path
from collections import defaultdict
from sklearn.metrics import f1_score, accuracy_score, recall_score, precision_score
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------- 配置参数 --------------------------
EXCEL_PATH = "/home/data/jy/GLIP/OUTPUTcentral_2017_val/密度对比.xlsx"
COCO_JSON_PATH = "/home/data/jy/GLIP/DATASET/selected_patches/coco_annotations-0820.json"
EPOCHS = 5000  # 训练轮数增加到5000
LEARNING_RATE = 1e-4  # 学习率
TEST_GOOD_COUNT = 20  # 测试集“好”样本数量
TEST_BAD_COUNT = 20  # 测试集“差”样本数量
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 测试集标红


# -----------------------------------------------------------------------------------

class WSIRegressionModel(nn.Module):
    """WSI预后回归模型"""

    def __init__(self, input_dim=2048, hidden_dim=512):
        super().__init__()
        self.fc_layers = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(hidden_dim, 1)
        )

    def forward(self, x):
        return self.fc_layers(x).squeeze()


def load_wsi_labels(excel_path):
    """从Excel加载WSI标签，按预后分组拆分（好/差）"""
    try:
        df_excel = pd.read_excel(excel_path)
    except Exception as e:
        raise RuntimeError(f"读取Excel失败: {str(e)}")

    # 按预后分组拆分样本（仅保留“好”和“差”，排除“不关心”）
    wsi_by_prognosis = {"好": [], "差": []}  # 格式：{预后: [(wsi_name, 标签值), ...]}
    for idx, row in df_excel.iterrows():
        try:
            wsi_name = str(row["filename"]).strip()
            prognosis = str(row["预后分组"]).strip()

            if prognosis == "好":
                wsi_by_prognosis["好"].append((wsi_name, 40))  # 好→标签40
            elif prognosis == "差":
                wsi_by_prognosis["差"].append((wsi_name, 0))  # 差→标签0
        except KeyError as e:
            raise KeyError(f"Excel中缺少必要列: {str(e)}")
        except Exception as e:
            print(f"处理Excel行 {idx} 时出错: {str(e)}, 跳过该行")

    # 统计各分组数量并校验是否满足测试集需求
    good_total = len(wsi_by_prognosis["好"])
    bad_total = len(wsi_by_prognosis["差"])
    print(f"加载到有效WSI:")
    print(f"- 预后'好'的WSI数量: {good_total} 个")
    print(f"- 预后'差'的WSI数量: {bad_total} 个")

    # 校验样本是否足够（需至少20个好+20个差）
    if good_total < TEST_GOOD_COUNT:
        raise ValueError(f"预后'好'的WSI仅 {good_total} 个，不足测试集需求（{TEST_GOOD_COUNT}个）")
    if bad_total < TEST_BAD_COUNT:
        raise ValueError(f"预后'差'的WSI仅 {bad_total} 个，不足测试集需求（{TEST_BAD_COUNT}个）")

    return wsi_by_prognosis, df_excel


def load_wsi_patches(coco_json_path):
    """从COCO JSON加载WSI对应的所有Patch路径"""
    try:
        with open(coco_json_path, "r", encoding="utf-8") as f:
            coco_data = json.load(f)
    except Exception as e:
        raise RuntimeError(f"读取COCO JSON失败: {str(e)}")

    images = coco_data.get("images", [])
    if not images:
        raise ValueError("COCO JSON中未找到images字段")

    # 提取WSI名（匹配"WSI名_patchxxx"格式）
    wsi_pattern = re.compile(r"^(.+?)_patch")
    patch_root = Path(coco_json_path).parent  # Patch与JSON同文件夹
    wsi_patch_paths = defaultdict(list)

    for img in images:
        patch_filename = img.get("file_name")
        if not patch_filename:
            continue

        wsi_match = wsi_pattern.match(patch_filename)
        if wsi_match:
            wsi_name = wsi_match.group(1)
            patch_abs_path = str(patch_root / patch_filename)
            wsi_patch_paths[wsi_name].append(patch_abs_path)

    print(f"成功加载 {len(wsi_patch_paths)} 个WSI的Patch路径")
    return wsi_patch_paths


def filter_samples_with_patch(wsi_list, wsi_patch_paths):
    """过滤出有有效Patch的样本（排除无Patch或Patch为空的WSI）"""
    valid_samples = []
    for wsi_name, label in wsi_list:
        if wsi_name not in wsi_patch_paths:
            print(f"警告: WSI {wsi_name} 无对应Patch，跳过")
            continue
        patch_paths = wsi_patch_paths[wsi_name]
        if len(patch_paths) == 0:
            print(f"警告: WSI {wsi_name} 无有效Patch，跳过")
            continue
        valid_samples.append((wsi_name, patch_paths, label))
    return valid_samples


def prepare_train_test_fixed(wsi_by_prognosis, wsi_patch_paths):
    """固定测试集为20个好+20个差，剩余为训练集"""
    np.random.seed(42)  # 固定随机种子，确保划分结果可复现

    # -------------------------- 处理“好”样本 --------------------------
    # 过滤有Patch的“好”样本
    good_wsi_list = wsi_by_prognosis["好"]
    good_valid = filter_samples_with_patch(good_wsi_list, wsi_patch_paths)
    print(f"\n预后'好'的有效样本（有Patch）: {len(good_valid)} 个")

    # 校验有效“好”样本是否足够
    if len(good_valid) < TEST_GOOD_COUNT:
        raise ValueError(f"预后'好'的有效样本仅 {len(good_valid)} 个，不足测试集需求（{TEST_GOOD_COUNT}个）")

    # 随机抽取20个“好”样本作为测试集，剩余作为训练集
    np.random.shuffle(good_valid)
    good_test = good_valid[:TEST_GOOD_COUNT]  # 前20个→测试集
    good_train = good_valid[TEST_GOOD_COUNT:]  # 剩余→训练集

    # -------------------------- 处理“差”样本 --------------------------
    # 过滤有Patch的“差”样本
    bad_wsi_list = wsi_by_prognosis["差"]
    bad_valid = filter_samples_with_patch(bad_wsi_list, wsi_patch_paths)
    print(f"预后'差'的有效样本（有Patch）: {len(bad_valid)} 个")

    # 校验有效“差”样本是否足够
    if len(bad_valid) < TEST_BAD_COUNT:
        raise ValueError(f"预后'差'的有效样本仅 {len(bad_valid)} 个，不足测试集需求（{TEST_BAD_COUNT}个）")

    # 随机抽取20个“差”样本作为测试集，剩余作为训练集
    np.random.shuffle(bad_valid)
    bad_test = bad_valid[:TEST_BAD_COUNT]  # 前20个→测试集
    bad_train = bad_valid[TEST_BAD_COUNT:]  # 剩余→训练集

    # -------------------------- 合并训练集和测试集 --------------------------
    train_samples = good_train + bad_train  # 训练集：剩余好 + 剩余差
    test_samples = good_test + bad_test  # 测试集：20好 + 20差

    # 打印划分结果
    print(f"\n数据划分完成（测试集固定20好+20差）:")
    print(f"- 训练集总样本: {len(train_samples)} 个WSI")
    print(f"  - 其中预后'好': {len(good_train)} 个")
    print(f"  - 其中预后'差': {len(bad_train)} 个")
    print(f"- 测试集总样本: {len(test_samples)} 个WSI")
    print(f"  - 其中预后'好': {len(good_test)} 个（固定）")
    print(f"  - 其中预后'差': {len(bad_test)} 个（固定）")

    # 计算训练集中的类别比例，用于后续权重计算
    total_train = len(train_samples)
    good_ratio = len(good_train) / total_train if total_train > 0 else 0
    bad_ratio = len(bad_train) / total_train if total_train > 0 else 0
    print(f"\n训练集类别比例: 好={good_ratio:.2%}, 差={bad_ratio:.2%}")

    return train_samples, test_samples, good_train, bad_train


def extract_wsi_features(samples, set_name="样本"):
    """提取WSI特征（基于Patch平均池化）"""
    # 配置设备（GPU优先）
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\n提取{set_name}特征，使用设备: {device}")

    # 图片预处理（匹配ResNet50预训练参数）
    patch_transform = transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225])
    ])

    # 加载预训练ResNet50（兼容新旧版本torchvision）
    try:
        cnn_model = models.resnet50(weights=models.ResNet50_Weights.IMAGENET1K_V1)
    except AttributeError:
        cnn_model = models.resnet50(pretrained=True)

    # 去掉最后一层全连接，保留特征提取部分
    cnn_model = torch.nn.Sequential(*list(cnn_model.children())[:-1])
    cnn_model = cnn_model.to(device)
    cnn_model.eval()

    # 单个Patch特征提取
    def extract_patch_feat(patch_path):
        try:
            img = Image.open(patch_path).convert("RGB")
            img_tensor = patch_transform(img).unsqueeze(0).to(device)
            with torch.no_grad():
                feat = cnn_model(img_tensor)  # (1, 2048, 1, 1)
            return feat.squeeze().cpu().numpy()  # 展平为(2048,)
        except Exception as e:
            print(f"提取Patch {patch_path} 特征失败: {str(e)}")
            return None

    # 批量提取WSI特征
    wsi_features = []
    wsi_labels = []
    valid_wsi_names = []
    # 记录样本类别用于后续加权
    sample_classes = []  # 0表示差，1表示好

    for idx, (wsi_name, patch_paths, label) in enumerate(samples):
        print(f"处理 {set_name} {idx + 1}/{len(samples)}: WSI={wsi_name}（Patch数: {len(patch_paths)}）")
        patch_feats = []

        # 提取该WSI所有有效Patch的特征
        for path in patch_paths:
            feat = extract_patch_feat(path)
            if feat is not None:
                patch_feats.append(feat)

        # 过滤无有效特征的WSI
        if len(patch_feats) == 0:
            print(f"警告: WSI {wsi_name} 无有效特征，跳过")
            continue

        # 特征聚合（平均池化）
        wsi_feat = torch.tensor(patch_feats).mean(dim=0)
        wsi_features.append(wsi_feat)
        wsi_labels.append(torch.tensor(label, dtype=torch.float32))
        valid_wsi_names.append(wsi_name)
        # 记录样本类别（0表示差，1表示好）
        sample_classes.append(1 if label == 40 else 0)

    # 转为Tensor并校验
    if len(wsi_features) == 0:
        raise RuntimeError(f"{set_name}无有效特征，无法继续")
    wsi_features = torch.stack(wsi_features).to(device)
    wsi_labels = torch.stack(wsi_labels).to(device)

    print(f"{set_name}特征提取完成: {len(wsi_features)} 个有效WSI，特征维度: {wsi_features.shape}")
    return wsi_features, wsi_labels, valid_wsi_names, device, sample_classes


def calculate_class_weights(sample_classes):
    """计算类别权重，解决类别不平衡问题"""
    # 统计每个类别的样本数量
    class_counts = np.bincount(sample_classes)
    total_samples = len(sample_classes)

    # 计算权重 (总样本数 / (类别数 * 该类样本数))
    weights = total_samples / (len(class_counts) * class_counts)

    # 归一化权重，使最大权重为1
    weights = weights / np.max(weights)

    print("\n计算得到类别权重:")
    for i, weight in enumerate(weights):
        class_name = "好" if i == 1 else "差"
        print(f"- 预后'{class_name}'的权重: {weight:.4f} (样本数: {class_counts[i]})")

    return torch.tensor(weights, dtype=torch.float32)


def train_model(train_feats, train_labels, device, class_weights, sample_classes):
    """训练回归模型（5000轮，带类别不平衡优化）"""
    # 初始化模型
    model = WSIRegressionModel().to(device)

    # 将类别权重移到设备上
    class_weights = class_weights.to(device)

    # 定义加权MSE损失函数
    def weighted_mse_loss(pred, target, weights, classes):
        # 计算基本MSE损失
        mse = nn.MSELoss(reduction='none')(pred, target)
        # 根据类别应用权重
        sample_weights = weights[classes]
        # 计算加权损失
        weighted_loss = torch.mean(mse * sample_weights)
        return weighted_loss

    # 优化器
    optimizer = optim.Adam(
        model.parameters(),
        lr=LEARNING_RATE,
        weight_decay=1e-5  # L2正则防过拟合
    )

    # 学习率调度器 - 随着训练进行降低学习率
    scheduler = optim.lr_scheduler.ReduceLROnPlateau(
        optimizer, mode='min', factor=0.5, patience=100, verbose=True
    )

    # 训练循环
    model.train()
    loss_history = []
    print(f"\n开始训练模型（共 {EPOCHS} 轮）...")

    # 将类别信息转换为Tensor并移到设备
    class_tensor = torch.tensor(sample_classes, dtype=torch.long).to(device)

    for epoch in range(EPOCHS):
        optimizer.zero_grad()

        # 前向传播
        outputs = model(train_feats)

        # 计算加权损失
        loss = weighted_mse_loss(outputs, train_labels, class_weights, class_tensor)
        loss_history.append(loss.item())

        # 反向传播+参数更新
        loss.backward()
        optimizer.step()

        # 每100轮打印日志（因轮数增加，调整打印频率）
        if (epoch + 1) % 100 == 0:
            print(f"Epoch [{epoch + 1}/{EPOCHS}] | Training Loss: {loss.item():.6f}")

        # 每50轮更新学习率
        if (epoch + 1) % 50 == 0:
            scheduler.step(loss)

    # 保存损失曲线
    plt.figure(figsize=(12, 6))
    plt.plot(range(1, EPOCHS + 1), loss_history, color="#1f77b4")
    plt.xlabel("Epoch", fontsize=12)
    plt.ylabel("Weighted MSE Loss", fontsize=12)
    plt.title(f"Training Loss (Test Set: 20 Good + 20 Bad, {EPOCHS} Epochs)", fontsize=14)
    plt.grid(alpha=0.3)
    plt.savefig("training_loss_20good20bad_weighted.png", dpi=300, bbox_inches="tight")
    print(f"\n训练损失曲线已保存为: training_loss_20good20bad_weighted.png")

    return model, loss_history


def evaluate_model(model, feats, labels, wsi_names, device, set_name):
    """评估模型"""
    model.eval()
    with torch.no_grad():
        # 回归预测→分类转换（阈值20：>20→好，≤20→差）
        pred_reg = model(feats)
        pred_cls = (pred_reg > 20).float()  # 预测分类（0=差，1=好）
        gt_cls = (labels == 40).float()  # 真实分类（0=差，1=好）

        # 转numpy计算指标
        pred_cls_np = pred_cls.cpu().numpy()
        gt_cls_np = gt_cls.cpu().numpy()
        pred_reg_np = pred_reg.cpu().numpy()

        # 计算核心指标
        accuracy = accuracy_score(gt_cls_np, pred_cls_np)
        precision = precision_score(gt_cls_np, pred_cls_np, zero_division=0)
        recall = recall_score(gt_cls_np, pred_cls_np, zero_division=0)
        f1 = f1_score(gt_cls_np, pred_cls_np, zero_division=0)

        # 统计测试集各类别预测情况
        if "测试集" in set_name:
            # 统计“好”样本预测正确数
            good_mask = gt_cls_np == 1
            good_correct = sum(pred_cls_np[good_mask] == gt_cls_np[good_mask])
            # 统计“差”样本预测正确数
            bad_mask = gt_cls_np == 0
            bad_correct = sum(pred_cls_np[bad_mask] == gt_cls_np[bad_mask])
            print(f"\n{set_name}类别预测详情:")
            print(f"- 预后'好'样本: {sum(good_mask)} 个，预测正确: {good_correct} 个，准确率: {good_correct / sum(good_mask):.2%}")
            print(f"- 预后'差'样本: {sum(bad_mask)} 个，预测正确: {bad_correct} 个，准确率: {bad_correct / sum(bad_mask):.2%}")

    # 打印评估结果
    print("\n" + "=" * 60)
    print(f"模型评估结果（{set_name}，总样本: {len(feats)}）")
    print("=" * 60)
    print(f"准确率（Accuracy）: {accuracy:.4f}")
    print(f"精确率（Precision）: {precision:.4f}")
    print(f"召回率（Recall）: {recall:.4f}")
    print(f"F1-score: {f1:.4f}")
    print("=" * 60)

    # 生成WSI→AI预后映射
    wsi_pred_map = {}
    for wsi_name, pred_val in zip(wsi_names, pred_reg_np):
        wsi_pred_map[wsi_name] = "好" if pred_val > 20 else "差"

    return wsi_pred_map, {
        "accuracy": accuracy, "precision": precision, "recall": recall, "f1": f1
    }


def update_excel(df_excel, excel_path, wsi_pred_map, test_wsi_names):
    """更新Excel：添加AI预后列 + 测试集标红"""
    # 1. 添加"AI预后"列（插入到"预后分组"后）
    if "AI预后" not in df_excel.columns:
        if "预后分组" in df_excel.columns:
            prog_col_idx = df_excel.columns.get_loc("预后分组")
            df_excel.insert(prog_col_idx + 1, "AI预后", "")
        else:
            df_excel["AI预后"] = ""
        print(f"\nExcel新增列: 'AI预后'")

    # 2. 填充AI预后结果
    for idx, row in df_excel.iterrows():
        try:
            wsi_name = str(row["filename"]).strip()
            if wsi_name in wsi_pred_map:
                df_excel.at[idx, "AI预后"] = wsi_pred_map[wsi_name]
        except Exception as e:
            print(f"处理Excel行 {idx} 出错: {str(e)}，跳过")

    # 3. 保存Excel用于标红
    df_excel.to_excel(excel_path, index=False, engine="openpyxl")
    print(f"AI预后结果已写入Excel: {excel_path}")

    # 4. 测试集行标红（匹配20好+20差的WSI名）
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # 找到"filename"列位置
        filename_col = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "filename":
                    filename_col = cell.column
                    break
            if filename_col:
                break

        if not filename_col:
            print("警告: 未找到'filename'列，无法标红测试集")
            wb.close()
            return

        # 标红测试集行
        test_wsi_set = set(test_wsi_names)
        marked_cnt = 0
        for row in range(2, ws.max_row + 1):  # 跳过表头
            cell_val = ws.cell(row=row, column=filename_col).value
            if isinstance(cell_val, str) and cell_val.strip() in test_wsi_set:
                # 整行标红
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = RED_FILL
                marked_cnt += 1

        # 保存标红结果
        wb.save(excel_path)
        print(f"测试集标红完成: 共 {marked_cnt} 行（20好+20差），Excel已保存")

    except Exception as e:
        print(f"测试集标红失败: {str(e)}")

    return df_excel


def main():
    try:
        # 步骤1: 加载数据（按预后分组拆分WSI + 加载Patch路径）
        print("=" * 60)
        print("步骤1: 加载WSI标签和Patch路径")
        print("=" * 60)
        wsi_by_prognosis, df_excel = load_wsi_labels(EXCEL_PATH)
        wsi_patch_paths = load_wsi_patches(COCO_JSON_PATH)

        # 步骤2: 划分训练集和测试集（测试集20好+20差）
        print("\n" + "=" * 60)
        print("步骤2: 划分训练集和测试集（测试集20好+20差）")
        print("=" * 60)
        train_samples, test_samples, good_train, bad_train = prepare_train_test_fixed(wsi_by_prognosis, wsi_patch_paths)

        # 步骤3: 提取训练集特征
        print("\n" + "=" * 60)
        print("步骤3: 提取训练集WSI特征")
        print("=" * 60)
        train_feats, train_labels, train_wsi, device, train_classes = extract_wsi_features(train_samples, "训练集")

        # 步骤4: 计算类别权重（解决不平衡问题）
        print("\n" + "=" * 60)
        print("步骤4: 计算类别权重（解决类别不平衡）")
        print("=" * 60)
        class_weights = calculate_class_weights(train_classes)

        # 步骤5: 训练模型（5000轮，带权重）
        print("\n" + "=" * 60)
        print("步骤5: 训练回归模型（5000轮）")
        print("=" * 60)
        model, loss_hist = train_model(train_feats, train_labels, device, class_weights, train_classes)

        # 步骤6: 提取测试集特征并评估
        print("\n" + "=" * 60)
        print("步骤6: 提取测试集特征并评估（20好+20差）")
        print("=" * 60)
        test_feats, test_labels, test_wsi, _, _ = extract_wsi_features(test_samples, "测试集")
        test_pred_map, test_metrics = evaluate_model(model, test_feats, test_labels, test_wsi, device, "测试集（20好+20差）")

        # 步骤7: 提取全量特征生成完整预测结果
        print("\n" + "=" * 60)
        print("步骤7: 提取全量数据特征（训练+测试）")
        print("=" * 60)
        all_samples = train_samples + test_samples
        all_feats, all_labels, all_wsi, _, _ = extract_wsi_features(all_samples, "全量数据")
        all_pred_map, all_metrics = evaluate_model(model, all_feats, all_labels, all_wsi, device, "全量数据")

        # 步骤8: 更新Excel（AI预后 + 测试集标红）
        print("\n" + "=" * 60)
        print("步骤8: 更新Excel并标红测试集")
        print("=" * 60)
        df_updated = update_excel(df_excel, EXCEL_PATH, all_pred_map, test_wsi)

        # 最终总结
        print("\n" + "=" * 60)
        print("所有操作完成！")
        print("=" * 60)
        print(f"1. 数据划分: 测试集固定20好+20差，训练集{len(train_samples)}个WSI")
        print(f"2. 模型训练: {EPOCHS}轮完成，使用类别加权损失解决不平衡问题")
        print(f"3. 损失曲线: 已保存为 training_loss_20good20bad_weighted.png")
        print(f"4. 测试集评估: F1={test_metrics['f1']:.4f}，准确率={test_metrics['accuracy']:.4f}")
        print(f"5. Excel更新: 新增'AI预后'列，{len(test_wsi)}个测试集行已标红")

    except Exception as e:
        print(f"\n操作失败: {str(e)}")
        raise


if __name__ == "__main__":
    main()
